"""
Report delivery service — combines daily pricing and staff performance into one email.
"""

from __future__ import annotations

import io
import logging
from datetime import date, timedelta

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.models.daily_pricing import DailyPricing
from app.models.hotel import Hotel
from app.models.user import User, UserRole
from app.services.email_service import send_email_with_attachment
from app.services.report import ReportService

logger = logging.getLogger(__name__)


def _is_valid_email(candidate: str | None) -> bool:
    if not candidate:
        return False
    value = candidate.strip()
    return "@" in value and "." in value.split("@")[-1]


async def collect_report_recipients(db: AsyncSession, hotel: Hotel) -> list[str]:
    """Collect recipients from owner email + active admin/supervisor user emails."""
    recipients: set[str] = set()

    if _is_valid_email(hotel.owner_email):
        recipients.add(hotel.owner_email.strip())

    users_stmt = select(User).where(
        User.hotel_id == hotel.id,
        User.is_active == True,
        User.role.in_([UserRole.ADMIN, UserRole.SUPERVISOR]),
    )
    users = (await db.execute(users_stmt)).scalars().all()

    for user in users:
        if _is_valid_email(user.email):
            recipients.add(user.email.strip().lower())
            continue

        # Backward-compatibility for old accounts where username was used as email.
        if _is_valid_email(user.username):
            recipients.add(user.username.strip().lower())

    return sorted(recipients)


async def fetch_pricing_rows(db: AsyncSession, hotel_id, report_date: date) -> list[DailyPricing]:
    stmt = (
        select(DailyPricing)
        .where(DailyPricing.hotel_id == hotel_id, DailyPricing.date == report_date)
        .order_by(DailyPricing.my_price.asc())
    )
    return (await db.execute(stmt)).scalars().all()


def _append_pricing_sheet(wb: Workbook, title: str, rows: list[DailyPricing]):
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    headers = ["اسم الفندق المنافس", "تاريخ التسعيرة", "سعر الفندق المنافس", "سعر فندقنا", "الفرق"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for p in rows:
        diff = float(p.my_price - p.competitor_price)
        if diff > 0:
            diff_text = f"أغلى بـ {diff}"
        elif diff < 0:
            diff_text = f"أرخص بـ {abs(diff)}"
        else:
            diff_text = "نفس السعر"

        ws.append([
            p.competitor_hotel_name,
            p.date.strftime("%Y-%m-%d"),
            float(p.competitor_price),
            float(p.my_price),
            diff_text,
        ])


def _append_staff_sheet(wb: Workbook, performance: dict):
    ws = wb.create_sheet("تقييم_الموظفين")
    ws.sheet_view.rightToLeft = True
    headers = [
        "الترتيب",
        "الموظف",
        "اسم المستخدم",
        "الدور",
        "حل الشكاوى",
        "تأكيد الحجوزات",
        "إجمالي العمليات",
        "متوسط زمن الحل (س)",
        "متوسط زمن الاعتماد (س)",
        "النقاط",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in performance.get("leaderboard", []):
        ws.append([
            row.get("rank", 0),
            row.get("full_name", "-"),
            row.get("username", "-"),
            row.get("role", "-"),
            row.get("complaints_resolved", 0),
            row.get("reservations_approved", 0),
            row.get("total_actions", 0),
            row.get("avg_resolution_hours", 0),
            row.get("avg_approval_hours", 0),
            row.get("score", 0),
        ])


def build_combined_xlsx(
    prices_today: list[DailyPricing],
    prices_yesterday: list[DailyPricing],
    performance: dict,
    report_date: date,
) -> bytes:
    wb = Workbook()
    first = wb.active
    wb.remove(first)

    _append_pricing_sheet(wb, f"اسعار_{report_date.strftime('%Y%m%d')}", prices_today)
    _append_pricing_sheet(wb, f"اسعار_{(report_date - timedelta(days=1)).strftime('%Y%m%d')}", prices_yesterday)
    _append_staff_sheet(wb, performance)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_combined_message(hotel_name: str, report_date: date, prices_today: list[DailyPricing], prices_yesterday: list[DailyPricing], performance: dict) -> str:
    yesterday = report_date - timedelta(days=1)

    def _line_for_price(p: DailyPricing) -> str:
        diff = float(p.my_price - p.competitor_price)
        if diff > 0:
            diff_mark = f"أغلى منا بـ {diff}"
        elif diff < 0:
            diff_mark = f"أرخص منا بـ {abs(diff)}"
        else:
            diff_mark = "نفس السعر"
        return f"• {p.competitor_hotel_name}: منافس {float(p.competitor_price)} | نحن {float(p.my_price)} ({diff_mark})"

    top = performance.get("leaderboard", [])[:3]

    lines = [
        "📊 تقرير يومي موحد من RAHATY",
        f"🏨 الفندق: {hotel_name}",
        f"📅 تاريخ التقرير: {report_date:%Y-%m-%d}",
        "",
        f"📈 أسعار اليوم ({report_date:%Y-%m-%d})",
    ]

    if prices_today:
        lines.extend([_line_for_price(p) for p in prices_today])
    else:
        lines.append("• لا توجد تسعيرات مسجلة لليوم")

    lines.extend([
        "",
        f"📉 أسعار الأمس ({yesterday:%Y-%m-%d})",
    ])
    if prices_yesterday:
        lines.extend([_line_for_price(p) for p in prices_yesterday])
    else:
        lines.append("• لا توجد تسعيرات مسجلة للأمس")

    summary = performance.get("summary", {})
    lines.extend([
        "",
        "🏆 ملخص تقييم الموظفين",
        f"• الموظفون النشطون: {summary.get('active_staff', 0)} من {summary.get('total_staff', 0)}",
        f"• حل الشكاوى: {summary.get('total_complaints_resolved', 0)}",
        f"• اعتماد الحجوزات: {summary.get('total_reservations_approved', 0)}",
        f"• متوسط زمن الحل: {summary.get('avg_response_hours', 0)} ساعة",
        f"• متوسط زمن الاعتماد: {summary.get('avg_approval_hours', 0)} ساعة",
        f"• معدل الرفض: {summary.get('rejection_rate', 0)}%",
        "",
        "🥇 أفضل الموظفين:",
    ])

    if top:
        for idx, row in enumerate(top, start=1):
            lines.append(
                f"{idx}) {row.get('full_name', '-')}: نقاط {row.get('score', 0)} | عمليات {row.get('total_actions', 0)}"
            )
    else:
        lines.append("لا توجد بيانات تقييم للفترة")

    lines.extend(["", "مرفق ملف Excel يحتوي على تفاصيل الأسعار لليومين وتقييم الموظفين."])
    return "\n".join(lines)


async def send_combined_pricing_staff_report(
    db: AsyncSession,
    hotel: Hotel,
    report_date: date,
    staff_days: int = 30,
) -> dict:
    """Send combined pricing + staff report to manager/supervisor recipients."""
    recipients = await collect_report_recipients(db, hotel)
    if not recipients:
        return {
            "success": False,
            "message": "No valid recipient emails found for admin/supervisor/owner",
            "recipients": [],
        }

    prices_today = await fetch_pricing_rows(db, hotel.id, report_date)
    prices_yesterday = await fetch_pricing_rows(db, hotel.id, report_date - timedelta(days=1))
    performance = await ReportService.generate_staff_performance(db, hotel.id, period_days=staff_days)

    attachment = build_combined_xlsx(prices_today, prices_yesterday, performance, report_date)
    body = build_combined_message(hotel.name, report_date, prices_today, prices_yesterday, performance)

    subject = f"تقرير موحد: الأسعار + تقييم الموظفين - {hotel.name} - {report_date:%Y-%m-%d}"
    sent_to = []
    failed_reasons: list[str] = []
    for email in recipients:
        try:
            await send_email_with_attachment(
                to_email=email,
                subject=subject,
                body_text=body,
                attachment_name=f"combined_report_{report_date:%Y%m%d}.xlsx",
                attachment_bytes=attachment,
            )
            sent_to.append(email)
        except Exception as ex:
            logger.error("❌ Failed sending combined report to %s: %s", email, ex)
            failed_reasons.append(f"{email}: {str(ex)}")

    return {
        "success": len(sent_to) > 0,
        "message": "Combined report sent" if sent_to else (failed_reasons[0] if failed_reasons else "Failed to send combined report"),
        "recipients": sent_to,
        "failed": failed_reasons,
        "total_recipients": len(sent_to),
    }
