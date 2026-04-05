"""
Daily Pricing API — manage daily prices for competitors vs our hotel.
"""

import uuid
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy.exc import IntegrityError

from app.database import get_db
from app.models.daily_pricing import DailyPricing
from app.schemas.daily_pricing import (
    DailyPricingCreate,
    DailyPricingUpdate,
    DailyPricingResponse,
    DailyPricingListResponse,
)

router = APIRouter(tags=["Daily Pricing"])


@router.get(
    "/hotels/{hotel_id}/daily-pricing",
    response_model=DailyPricingListResponse,
)
async def list_daily_pricing(
    hotel_id: uuid.UUID,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    skip: int = 0,
    limit: int = 50,
    db: AsyncSession = Depends(get_db),
):
    """List daily pricing with optional date filters."""
    query = select(DailyPricing).where(DailyPricing.hotel_id == hotel_id)

    if from_date:
        query = query.where(DailyPricing.date >= from_date)
    if to_date:
        query = query.where(DailyPricing.date <= to_date)

    # Count
    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar() or 0

    # Fetch
    query = query.order_by(DailyPricing.date.desc(), DailyPricing.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(query)
    prices = result.scalars().all()

    return DailyPricingListResponse(
        items=[DailyPricingResponse.model_validate(p) for p in prices],
        total=total,
    )


@router.post(
    "/hotels/{hotel_id}/daily-pricing",
    response_model=DailyPricingResponse,
    status_code=201,
)
async def create_daily_pricing(
    hotel_id: uuid.UUID,
    data: DailyPricingCreate,
    db: AsyncSession = Depends(get_db),
):
    """Create a daily pricing entry."""
    pricing_date = data.date if data.date else date.today()
    
    pricing = DailyPricing(
        hotel_id=hotel_id,
        competitor_hotel_name=data.competitor_hotel_name,
        date=pricing_date,
        my_price=data.my_price,
        competitor_price=data.competitor_price
    )
    
    try:
        db.add(pricing)
        await db.commit()
        await db.refresh(pricing)
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=400, 
            detail=f"Pricing for {data.competitor_hotel_name} on {pricing_date} already exists."
        )

    return DailyPricingResponse.model_validate(pricing)


@router.put(
    "/hotels/{hotel_id}/daily-pricing/{pricing_id}",
    response_model=DailyPricingResponse,
)
async def update_daily_pricing(
    hotel_id: uuid.UUID,
    pricing_id: uuid.UUID,
    data: DailyPricingUpdate,
    db: AsyncSession = Depends(get_db),
):
    """Update a daily pricing entry."""
    stmt = select(DailyPricing).where(
        DailyPricing.id == pricing_id,
        DailyPricing.hotel_id == hotel_id,
    )
    result = await db.execute(stmt)
    pricing = result.scalar_one_or_none()
    
    if not pricing:
        raise HTTPException(status_code=404, detail="Daily pricing entry not found")

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(pricing, field, value)

    try:
        await db.commit()
        await db.refresh(pricing)
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=400, detail="Update causes duplicate entry")

    return DailyPricingResponse.model_validate(pricing)


@router.delete(
    "/hotels/{hotel_id}/daily-pricing/{pricing_id}",
    response_model=dict,
)
async def delete_daily_pricing(
    hotel_id: uuid.UUID,
    pricing_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    """Delete a daily pricing entry."""
    stmt = select(DailyPricing).where(
        DailyPricing.id == pricing_id,
        DailyPricing.hotel_id == hotel_id,
    )
    result = await db.execute(stmt)
    pricing = result.scalar_one_or_none()
    
    if not pricing:
        raise HTTPException(status_code=404, detail="Daily pricing entry not found")

    await db.delete(pricing)
    await db.commit()

    return {"success": True, "message": "Pricing entry deleted"}


@router.get(
    "/hotels/{hotel_id}/daily-pricing/export",
    tags=["Reports"],
)
async def export_daily_pricing(
    hotel_id: uuid.UUID,
    report_date: date = Query(..., alias="date"),
    db: AsyncSession = Depends(get_db),
):
    """Export daily pricing for a specific date as CSV."""
    import io
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    stmt = select(DailyPricing).where(
        DailyPricing.hotel_id == hotel_id,
        DailyPricing.date == report_date
    ).order_by(DailyPricing.my_price.asc())

    result = await db.execute(stmt)
    prices = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "التسعير اليومي"
    ws.sheet_view.rightToLeft = True

    headers = ["اسم الفندق المنافس", "تاريخ التسعيرة", "سعر الفندق المنافس", "سعر فندقنا", "الفرق"]
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for p in prices:
        diff = float(p.my_price - p.competitor_price)
        if diff > 0:
            diff_text = f"أغلى بـ {diff}"
        elif diff < 0:
            diff_text = f"أرخص بـ {abs(diff)}"
        else:
            diff_text = "نفس السعر"
            
        ws.append([
            p.competitor_hotel_name,
            p.date.strftime("%Y-%m-%d"),
            float(p.competitor_price),
            float(p.my_price),
            diff_text
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="daily_pricing_{report_date.strftime("%Y%m%d")}.xlsx"'}
    )


@router.post(
    "/hotels/{hotel_id}/daily-pricing/send-report",
    response_model=dict,
)
async def send_daily_pricing_report(
    hotel_id: uuid.UUID,
    report_date: date = Query(..., alias="date"),
    db: AsyncSession = Depends(get_db),
):
    """Manually send the daily pricing report to the hotel owner."""
    from app.models.hotel import Hotel
    from app.whatsapp.client import whatsapp_client

    hotel = await db.get(Hotel, hotel_id)
    if not hotel:
        raise HTTPException(404, "Hotel not found")

    stmt = select(DailyPricing).where(
        DailyPricing.hotel_id == hotel_id,
        DailyPricing.date == report_date
    ).order_by(DailyPricing.my_price.asc())
    
    result = await db.execute(stmt)
    prices = result.scalars().all()

    if not prices:
        return {"success": False, "message": "No pricing found for this date"}

    msg_lines = [
        f"📊 *تقرير أسعار المنافسين اليومي*",
        f"🏨 الفندق: *{hotel.name}*",
        f"📅 التاريخ: {report_date.strftime('%Y-%m-%d')}\n"
    ]

    for p in prices:
        diff = float(p.my_price - p.competitor_price)
        if diff > 0:
            diff_mark = f"أغلى منا بـ {diff}" # wait, if my_price is larger, we are more expensive
            diff_mark = f"أرخص منا بـ {diff}"
        elif diff < 0:
            diff_mark = f"أغلى منا بـ {abs(diff)}"
        else:
            diff_mark = "نفس السعر"
            
        msg_lines.append(f"• *{p.competitor_hotel_name}*: {float(p.competitor_price)} ريال (نحن: {float(p.my_price)} ريال) ⟵ {diff_mark}")

    msg_lines.append("\nتم إصدار التقرير من نظام إدارة الفنادق الذكي ✨")
    message = "\n".join(msg_lines)

    xlsx_bytes = None
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from app.services.email_service import send_email_with_attachment

        wb = Workbook()
        ws = wb.active
        ws.title = "التسعير اليومي"
        ws.sheet_view.rightToLeft = True

        headers = ["اسم الفندق المنافس", "تاريخ التسعيرة", "سعر الفندق المنافس", "سعر فندقنا", "الفرق"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for p in prices:
            diff = float(p.my_price - p.competitor_price)
            if diff > 0:
                diff_text = f"أغلى بـ {diff}"
            elif diff < 0:
                diff_text = f"أرخص بـ {abs(diff)}"
            else:
                diff_text = "نفس السعر"
                
            ws.append([
                p.competitor_hotel_name,
                p.date.strftime("%Y-%m-%d"),
                float(p.competitor_price),
                float(p.my_price),
                diff_text
            ])

        output = io.BytesIO()
        wb.save(output)
        xlsx_bytes = output.getvalue()
        output.close()
        wb.close()
        import logging
        logging.getLogger(__name__).info(f"📊 Excel report generated: {len(xlsx_bytes)} bytes")
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"❌ Failed to generate excel report for {hotel.name}: {e}")
        xlsx_bytes = None

    # Send Email if owner_email exists
    email_success = False
    if hotel.owner_email and xlsx_bytes:
        try:
            # AI generated email body
            ai_body = "مرفق طيه تقرير أسعار المنافسين اليومي بصيغة نظام إدارة الفنادق الذكي.\n\n" + message
            try:
                from app.config import get_settings
                from openai import AsyncOpenAI
                settings_ai = get_settings()
                if settings_ai.OPENAI_API_KEY:
                    ai_client = AsyncOpenAI(api_key=settings_ai.OPENAI_API_KEY)
                    ai_prompt = (
                        f"أنت مساعد افتراضي ذكي يعمل في 'نظام إدارة الفنادق الذكي'. "
                        f"اكتب رسالة بريد إلكتروني رسمية، مرحبة، وقصيرة إلى مدير فندق '{hotel.name}'.\n\n"
                        f"قم بالترحيب به بصفته مدير الفندق، ولخص له أسعار المنافسين اليوم بناءً على هذا التقرير النصي:\n"
                        f"{message}\n\n"
                        f"أخبره بلطف أن التقرير التفصيلي لمعرفة كافة الفروقات مرفق كملف إكسل مع هذا الإيميل لتسهيل قراءته. "
                        f"تحدث بلغة عربية فصحى احترافية ولا تضف أرقاماً من كيسك. كن مباشراً ولبقاً."
                    )
                    ai_response = await ai_client.chat.completions.create(
                        model=settings_ai.OPENAI_MODEL,
                        messages=[{"role": "user", "content": ai_prompt}],
                        max_tokens=1500,
                    )
                    ai_body = ai_response.choices[0].message.content
                    # Append hidden marker for AI agent identification
                    ai_body += f"\n\n[HID:{hotel.id}]"
            except Exception as ai_e:
                pass

            await send_email_with_attachment(
                to_email=hotel.owner_email,
                subject=f"تقرير أسعار المنافسين - {hotel.name} - {report_date.strftime('%Y-%m-%d')}",
                body_text=ai_body,
                attachment_name=f"daily_pricing_{report_date.strftime('%Y%m%d')}.xlsx",
                attachment_bytes=xlsx_bytes
            )
            email_success = True
        except ValueError as ve:
            import logging
            logging.getLogger(__name__).error(f"Email config error: {ve}")
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"Email sending failed: {e}")
            raise HTTPException(400, f"فشل إرسال الإيميل. تأكد من الباسوورد والإيميل: {str(e)}")

    # Use telegram if preferred, or whatsapp
    hotel_settings = hotel.settings or {}
    tg_token = hotel.telegram_bot_token or hotel_settings.get("telegram_bot_token")
    
    from app.config import get_settings
    settings_app = get_settings()
    tg_tok = tg_token or settings_app.TELEGRAM_BOT_TOKEN
    wa_id = hotel.whatsapp_phone_number_id or settings_app.WHATSAPP_PHONE_NUMBER_ID

    send_result = None

    # Check length to guess if it's telegram chat ID or phone number
    is_mostly_telegram = hotel.owner_whatsapp.startswith("tg_") or (hotel.owner_whatsapp.isdigit() and len(hotel.owner_whatsapp) <= 10)

    if is_mostly_telegram and tg_tok:
        send_result = await whatsapp_client.send_telegram_message(
            bot_token=tg_tok,
            chat_id=hotel.owner_whatsapp,
            message=message
        )
    elif wa_id and hotel.owner_whatsapp:
        send_result = await whatsapp_client.send_text_message(
            phone_number_id=wa_id,
            to=hotel.owner_whatsapp,
            message=message,
            api_token=hotel.whatsapp_api_token
        )
    elif tg_tok and hotel.owner_whatsapp:
         # Fallback to telegram if WA is missing entirely
         send_result = await whatsapp_client.send_telegram_message(
            bot_token=tg_tok,
            chat_id=hotel.owner_whatsapp,
            message=message
        )

    if send_result and "error" in send_result and not email_success:
        import logging
        logging.getLogger(__name__).error(f"❌ Both WhatsApp and Email failed for {hotel.name}. WA Error: {send_result['error']}")
        raise HTTPException(400, f"فشل الإرسال بكل القنوات: {send_result['error']}")

    return {
        "success": True, 
        "message": "تم إرسال التقرير بنجاح",
        "email_sent": email_success,
        "whatsapp_sent": not (send_result and "error" in send_result)
    }
